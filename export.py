import os
import sys

sys.path.insert(0, '/home/welkere/python')
sys.path.insert(0, '/home/welkere/python/pmh_img')

os.environ['DJANGO_SETTINGS_MODULE'] = 'pmh_img.settings'

from images.models import Image, TopicPage
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

wb = Workbook()
ws = wb.worksheets[0]
ws.title = 'PMH Anatomy Images'
fn = r'pmh_imgs.xlsx'

col_names = ['Term',
            'MeSH',
            'Fig1 original filename',
            'Fig1 PMH filename',
            'Fig1 source: Credit line',
            'Fig1 source: Displayed credit',
            'Fig1 source URL',
            'Fig1 alt-text',
            'Fig1 image-associated caption',
            'Fig1 topic page caption',
            'Fig2 original filename',
            'Fig2 PMH filename',
            'Fig2 source: Credit line',
            'Fig2 source: Displayed credit',
            'Fig2 source URL',
            'Fig2 alt-text',
            'Fig2 image-associated caption',
            'Fig2 topic page caption',
            'Fig3 original filename',
            'Fig3 PMH filename',
            'Fig3 source: Credit line',
            'Fig3 source: Displayed credit',
            'Fig3 source URL',
            'Fig3 alt-text',
            'Fig3 image-associated caption',
            'Fig3 topic page caption',
            'Snippet: ID for the definition',
            'Snippet: content of the definition',
            'Snippet source',
            'URL for snippet source',
            'Teaser title',
            'Teaser text ',
            'Teaser PMHID if more to read',
            'Source for teaser',
            'URL for teaser source',
            'Parent terms',
            'Child terms ',
            'Synonyms',
            'Reldoc-see-also',
            'Reldoc-see-also sources',
            'MeSH for related conditions',
            'MeSH for terms to know']


for idx, name in enumerate(col_names):
    col = get_column_letter((idx+1))
    ws.cell('%s%s' % (col, '1')).value = name

topics = TopicPage.objects.order_by('topic')
for idx, topic in enumerate(topics):
    #first row is already headers, and w/0 index
    row = idx + 2
    ws.cell('A%s' % (row,)).value = topic.topic
    ws.cell('B%s' % (row,)).value = topic.mesh_codes

    ws.cell('C%s' % (row,)).value = topic.image.image.name
    ws.cell('D%s' % (row,)).value = topic.image.name
    ws.cell('E%s' % (row,)).value = topic.image.orig_figure_source
    ws.cell('F%s' % (row,)).value = topic.image.pmh_figure_source
    ws.cell('G%s' % (row,)).value = topic.image.source_url
    ws.cell('H%s' % (row,)).value = topic.image.alt_text
    ws.cell('I%s' % (row,)).value = topic.image.caption
    ws.cell('J%s' % (row,)).value = topic.image.pmh_caption

ws = wb.create_sheet()

wb.save(filename = fn)
